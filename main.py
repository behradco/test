import requests
from openpyxl import Workbook
def conect_Api_Gather_info():
    contact = {
        "contact": {
            'DisplayName': 'معمارنژاد-قم',
            "ContactType": "DontCare",
        },
        "From": 0,
        "Limit": 6000
    }
    req = requests.post("https://app.didar.me/api/deal/search?apikey=g170pxn6spxqrb4pac65ouiawf7ifyxv", json=contact)
    json_response = req.json()

    repository = json_response["Response"]["List"]
    # print(repository)
    total_info_list = []  # total list for keeping a list of dictionary of items
    for items in repository:
        info_dict = {}
        if items["Status"]=='Won':
            temp_arr = []
            if items["Contact"] != None:
                info_dict["DisplayName"] = items["Contact"]["DisplayName"]
                info_dict["MobilePhone"]=items["Contact"]["MobilePhone"]
            else:
                info_dict["DisplayName"] = "No Name"
                info_dict["MobilePhone"]="No Number"
            info_dict["Date"] = items['ChangeToWonTime']
            if len(items["Invoice"]["Items"]) != 0:
                temp_arr = []
                for K in items["Invoice"]["Items"]:
                    temp_dict = {}
                    temp_dict["OriginalTitle"] = K["OriginalTitle"]
                    temp_dict["Quantity"] = K["Quantity"]
                    temp_arr.append(temp_dict)

            info_dict["Products"]=temp_arr

            total_info_list.append(info_dict)
            print(total_info_list)
    return(total_info_list)
def gregorian_to_jalali(gy, gm, gd):
    g_d_m = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334]
    if (gm > 2):
        gy2 = gy + 1
    else:
        gy2 = gy
    days = 355666 + (365 * gy) + ((gy2 + 3) // 4) - ((gy2 + 99) // 100) + ((gy2 + 399) // 400) + gd + g_d_m[gm - 1]
    jy = -1595 + (33 * (days // 12053))
    days %= 12053
    jy += 4 * (days // 1461)
    days %= 1461
    if (days > 365):
        jy += (days - 1) // 365
        days = (days - 1) % 365
    if (days < 186):
        jm = 1 + (days // 31)
        jd = 1 + (days % 31)
    else:
        jm = 7 + ((days - 186) // 30)
        jd = 1 + ((days - 186) % 30)
    return [jy, jm, jd]
def convert_date(jy, jm, jd):
    jy += 1595
    days = -355668 + (365 * jy) + ((jy // 33) * 8) + (((jy % 33) + 3) // 4) + jd
    if (jm < 7):
        days += (jm - 1) * 31
    else:
        days += ((jm - 7) * 30) + 186
    gy = 400 * (days // 146097)
    days %= 146097
    if (days > 36524):
        days -= 1
        gy += 100 * (days // 36524)
        days %= 36524
        if (days >= 365):
            days += 1
    gy += 4 * (days // 1461)
    days %= 1461
    if (days > 365):
        gy += ((days - 1) // 365)
        days = (days - 1) % 365
    gd = days + 1
    if ((gy % 4 == 0 and gy % 100 != 0) or (gy % 400 == 0)):
        kab = 29
    else:
        kab = 28
    sal_a = [0, 31, kab, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    gm = 0
    while (gm < 13 and gd > sal_a[gm]):
        gd -= sal_a[gm]
        gm += 1

    return [gy, gm, gd]
def systemdate():
    import datetime
    x = datetime.datetime.now()
    gd = x.day
    gm = x.month
    gy = x.year
    y, m, d = gregorian_to_jalali(gy, gm, gd)
    return(y,m,d)
def desire_date(jal_y, jal_m, jal_d):
    if jal_d == 15 and jal_m == 1:
        jal_d_pass = 29
        jal_m_pass = 12
        jal_y_pass = jal_y - 1
    elif jal_d < 15 and jal_m == 1:
        temp = 15 - jal_d
        jal_d_pass = 29 - temp
        jal_m_pass = 12
        jal_y_pass = jal_y - 1
    elif jal_d < 15 and jal_m <= 7:
        temp = 15 - jal_d
        jal_d_pass = 31 - temp
        jal_m_pass = jal_m - 1
        jal_y_pass = jal_y
    elif jal_d < 15 and jal_m > 7:
        temp = 15 - jal_d
        jal_d_pass = 30 - temp
        jal_m_pass = jal_m - 1
        jal_y_pass = jal_y
    else:
        jal_d_pass = jal_d-15
        jal_m_pass = jal_m
        jal_y_pass = jal_y

    return (jal_y_pass, jal_m_pass, jal_d_pass)
def split_date(d):
    splited_date=(d.split("T"))[0]
    year_mounth_day = splited_date.split("-")
    return year_mounth_day

def extract_data(ger_y, ger_m, ger_d, extracted_list):
    temp_array=[]
    for i in extracted_list:
        temp_date_list=split_date(i["Date"])
        if int(temp_date_list[2])==int(ger_d):
            # print("yes 1")
            if int(temp_date_list[1])==int(ger_m):
                # print("yes 2")
                if int(temp_date_list[0]) == int(ger_y):
                    # print("yes 3")
                    temp_array.append(i)
                    # print(temp_array)
    return(temp_array)
def create_info(array,y, m, d, sys_y,sys_m, sys_d ):
    name=[]
    mobilephone=[]
    sms_date=[]
    win_sms=[]
    ups_number=[]
    bat_number=[]

    for i in array:
        upsnumber=0
        batnumber=0
        sms=str(sys_y) + "/" + str(sys_m) + "/" + str(sys_d)
        win = str(y) + "/" + str(m) + "/" + str(d)
        name.append(i["DisplayName"])
        mobilephone.append(i["MobilePhone"])
        sms_date.append(sms)
        win_sms.append(win)
        # print(i["Products"][1])
        for j in i["Products"]:
            if "یو پی اس" in j['OriginalTitle']:
                upsnumber = +int(j['Quantity'])
            if "باتری" in j['OriginalTitle']:
                batnumber = +int(j['Quantity'])
        ups_number.append(upsnumber)
        bat_number.append(batnumber)
    print(ups_number, bat_number)
    return(name, mobilephone,sms_date, win_sms, ups_number, bat_number)
def create_excelfile(name, mobilename, sms_date, win_date, ups, bat):
    wb = Workbook()
    # wb.create_sheet("Sheet_one")
    ws = wb['Sheet']
    ws['A1'] = 'نام شرکت'
    ws['B1'] = 'تاریخ موفق معامله'
    ws['C1'] = 'تاریخ ارسال SMS'
    ws['D1'] = 'شماره تلفن'
    ws['E1'] = 'تعداد یو پی اس'
    ws['F1'] = 'تعداد باتری'
    ws['G1']='متن اس ام اس'
    str1=f'سلام جناب آقای/ خانم {name}  با توجه به خرید {ups} عدد دستگاه یو پی اس و {bat} عدد باتری در تاریخ {win_date} خواهشمندیم جهت ارتقای سطح خدمات و آگاهی از نظرات ارزشمندتان فرم رضایت سنجی زیر را تکمیل نمایید. '
    row_start = 2  # start below the header row 2
    col_start = 1  # starts from column B
    print(f"str is {str1}")
    for i in range(0, len(name)):
        ws.cell(row_start + i, col_start).value = name[i]
        ws.cell(row_start + i, col_start + 1).value =win_date[i]
        ws.cell(row_start + i, col_start+2).value =sms_date[i]
        ws.cell(row_start + i, col_start + 3).value =  mobilename[i]
        ws.cell(row_start + i, col_start + 4).value = str(ups[i])
        ws.cell(row_start + i, col_start + 5).value = str(bat[i])
        ws.cell(row_start + i, col_start + 6).value = str1


    wb.save('book_eg.xlsx')

def main():
    # sys_y,sys_m, sys_d=systemdate()
    sys_y=1401
    sys_m=5
    sys_d=27
    y, m, d= desire_date(sys_y, sys_m, sys_d)
    print(y, m, d)
    ger_y, ger_m, ger_d=convert_date(y,m, d)
    print(ger_y, ger_m, ger_d)
    extracted_list=conect_Api_Gather_info()
    print(extracted_list)
    array=extract_data(ger_y, ger_m, ger_d, extracted_list)
    print(f"array is {array}")
    name, mobilename, sms_date, win_date, ups_number, battery_number=create_info(array,y, m, d, sys_y,sys_m, sys_d, )
    create_excelfile(name, mobilename, sms_date, win_date,ups_number, battery_number)

main()





